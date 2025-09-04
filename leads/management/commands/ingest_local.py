from __future__ import annotations
import csv
import hashlib
import re
from pathlib import Path
from urllib.parse import urlparse
from datetime import datetime
from django.utils import timezone

from django.core.management.base import BaseCommand
from django.db import transaction, IntegrityError
from openpyxl import load_workbook

from leads.models import State, City, Category, Source, SourceFile, Lead
from datetime import date


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def parse_city_state_from_filename(name: str):
    m = re.search(r"_in_(.*)\.csv$", name)
    if not m:
        return None, None
    tail = m.group(1)
    parts = tail.split('_')
    if len(parts) < 2:
        return None, None
    state = parts[-1]
    city = ' '.join(parts[:-1])
    return city.replace('-', ' ').strip(), state.replace('-', ' ').strip()


def normalize_domain(website: str | None, email: str | None) -> str | None:
    def clean(d: str) -> str:
        d = d.lower().strip()
        d = d.split('@')[-1] if '@' in d else d
        if d.startswith('http://') or d.startswith('https://'):
            d = urlparse(d).netloc
        d = d.split('/')[0]
        if d.startswith('www.'):
            d = d[4:]
        return d
    if website:
        try:
            return clean(website)
        except Exception:
            pass
    if email:
        try:
            return clean(email)
        except Exception:
            pass
    return None


def pick(row: dict, keys: list[str]):
    for k in keys:
        if k in row and row[k]:
            return row[k]
    return None


def clip(value: str | None, n: int) -> str | None:
    if value is None:
        return None
    s = str(value)
    return s[:n]


class Command(BaseCommand):
    help = 'Ingest CSVs from a local folder into the database.'

    def add_arguments(self, parser):
        parser.add_argument('--root', type=str, default='data/USA Database Business Leads')
        parser.add_argument('--source-name', dest='source_name', type=str, default='local')
        parser.add_argument('--glob', type=str, default='all', help='all (CSV+XLSX) or rglob pattern, or comma-separated patterns')
        parser.add_argument('--limit', type=int, default=None, help='Ingest at most N files (for testing)')

    def handle(self, *args, **opts):
        root = Path(opts['root'])
        source_name = opts.get('source_name', 'local')
        source, _ = Source.objects.get_or_create(name=source_name, defaults={'type': 'local_folder', 'root_path': str(root)})

        # Expand file patterns
        if opts['glob'] == 'all':
            patterns = ['**/*.csv', '**/*.CSV', '**/*.xlsx', '**/*.XLSX']
        else:
            patterns = [p.strip() for p in opts['glob'].split(',') if p.strip()]
        seen = {}
        file_paths = []
        for pat in patterns:
            for p in root.rglob(pat):
                sp = str(p)
                if sp not in seen:
                    seen[sp] = True
                    file_paths.append(p)
        file_paths.sort(key=lambda p: str(p))
        if opts.get('limit'):
            file_paths = file_paths[: int(opts['limit'])]
        self.stdout.write(f"Found {len(file_paths)} CSV files to consider.")
        for i, path in enumerate(sorted(file_paths)):
            try:
                self.ingest_file(source, path)
            except Exception as e:
                self.stderr.write(f"Error processing {path}: {e}")
        self.stdout.write(self.style.SUCCESS('Ingestion complete.'))

    def ingest_file(self, source: Source, path: Path):
        sha = file_sha256(path)
        stat = path.stat()

        # Category = parent directory name relative to root (handles nested datasets)
        try:
            rel = path.relative_to(Path(source.root_path))
            parts = rel.parts
            category_name = parts[-2] if len(parts) >= 2 else path.parent.name
        except Exception:
            category_name = path.parent.name
        city_name, state_name = parse_city_state_from_filename(path.name)
        category, _ = Category.objects.get_or_create(name=category_name)
        state = None
        city = None
        if state_name:
            state, _ = State.objects.get_or_create(name=state_name)
        if state and city_name:
            city, _ = City.objects.get_or_create(name=city_name.replace('_', ' '), state=state)

        sf, created = SourceFile.objects.get_or_create(
            source=source, path=str(path),
            defaults={'hash': sha, 'size': stat.st_size, 'modified_time': timezone.make_aware(datetime.fromtimestamp(stat.st_mtime)), 'category': category, 'state': state, 'city': city}
        )
        if not created and sf.hash == sha:
            # No change
            return

        # Update metadata
        sf.hash = sha
        sf.size = stat.st_size
        sf.modified_time = timezone.make_aware(datetime.fromtimestamp(stat.st_mtime))
        sf.category = category
        sf.state = state
        sf.city = city
        sf.save()

        def iter_rows_from_csv(p: Path):
            with p.open(newline='', encoding='utf-8-sig', errors='ignore') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    yield row

        def iter_rows_from_xlsx(p: Path):
            wb = load_workbook(filename=str(p), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header = None
            for r in ws.iter_rows(values_only=True):
                if header is None:
                    header = [(str(c).strip() if c is not None else '') for c in r]
                    # ensure unique keys
                    seen = {}
                    for i, h in enumerate(header):
                        if not h:
                            h = f'col_{i+1}'
                        if h in seen:
                            seen[h] += 1
                            h = f"{h}_{seen[h]}"
                        else:
                            seen[h] = 1
                        header[i] = h
                    continue
                vals = list(r)
                row = {}
                for i, h in enumerate(header):
                    v = vals[i] if i < len(vals) else None
                    row[h] = '' if v is None else v
                yield row

        ext = path.suffix.lower()
        if ext == '.csv':
            row_iter = iter_rows_from_csv(path)
        elif ext == '.xlsx':
            row_iter = iter_rows_from_xlsx(path)
        else:
            return

        count = 0
        with transaction.atomic():
            for row in row_iter:
                business_name = clip(pick(row, ['Name', 'Company', 'Business Name', 'Full Name']) or 'Unknown', 255)
                website = clip(pick(row, ['Website', 'Company Website']), 255)
                email = clip(pick(row, ['Company Email', 'Work Email #1', 'Direct Email #1']), 255)
                phone = clip(pick(row, ['Phone', 'Company Phone', 'Phone #1']), 100)
                address = pick(row, ['Address', 'Location'])
                domain = normalize_domain(website, email)

                # Attempt to enrich city/state from row if file-level parsing failed
                row_city, row_state = None, None
                q_val = row.get('Query') or ''
                if q_val:
                    q_val = q_val.replace('_', ' ')
                    m = re.search(r" in (.*)", q_val)
                    if m:
                        tail = m.group(1).strip()
                        parts = tail.split()
                        if len(parts) >= 2:
                            row_state = parts[-1]
                            row_city = ' '.join(parts[:-1])
                if not row_city and row.get('City'):
                    row_city = row.get('City')
                if not row_state and row.get('State'):
                    row_state = row.get('State')

                # Resolve city/state objects, fallback to file-level
                st = state
                ct = city
                if row_state:
                    st, _ = State.objects.get_or_create(name=str(row_state))
                if st and row_city:
                    ct, _ = City.objects.get_or_create(name=str(row_city), state=st)

                # Simple quality score heuristic
                score = 0
                if email:
                    score += 40
                if website:
                    score += 30
                if phone:
                    score += 20
                if row.get('Rating'):
                    try:
                        score += min(int(float(row['Rating']) * 2), 10)
                    except Exception:
                        pass

                # Upsert by email or domain+geo
                obj = None
                if email:
                    obj = Lead.objects.filter(email__iexact=email).first()
                if not obj and domain and city and state:
                    obj = Lead.objects.filter(domain__iexact=domain, city=city, state=state).first()

                if obj:
                    # Update minimal fields and last_seen
                    obj.business_name = business_name
                    obj.website = website or obj.website
                    obj.email = email or obj.email
                    obj.phone = phone or obj.phone
                    obj.address = address or obj.address
                    obj.category = obj.category or category
                    obj.state = obj.state or st
                    obj.city = obj.city or ct
                    obj.domain = obj.domain or domain
                    obj.quality_score = max(obj.quality_score, score)
                    obj.source_file = sf
                    obj.save()
                else:
                    try:
                        # Ensure JSON serializable 'extra'
                        safe_extra = {}
                        for k, v in row.items():
                            if isinstance(v, (datetime, date)):
                                safe_extra[k] = v.isoformat()
                            else:
                                safe_extra[k] = v
                        Lead.objects.create(
                                business_name=business_name,
                                website=website,
                                email=email,
                                phone=phone,
                                address=address,
                                category=category,
                                state=st,
                                city=ct,
                                domain=domain,
                                quality_score=score,
                                extra=safe_extra,
                                source_file=sf,
                        )
                    except IntegrityError:
                        # If unique constraint triggers, fetch existing and update
                        existing = None
                        if domain and st and ct:
                            existing = Lead.objects.filter(domain__iexact=domain, state=st, city=ct).first()
                        if not existing and email:
                            existing = Lead.objects.filter(email__iexact=email).first()
                        if existing:
                            existing.business_name = business_name or existing.business_name
                            existing.website = website or existing.website
                            existing.email = email or existing.email
                            existing.phone = phone or existing.phone
                            existing.address = address or existing.address
                            existing.category = existing.category or category
                            existing.state = existing.state or st
                            existing.city = existing.city or ct
                            existing.domain = existing.domain or domain
                            existing.quality_score = max(existing.quality_score, score)
                            existing.source_file = sf
                            existing.save()
                count += 1

        sf.row_count = count
        sf.last_ingested_at = timezone.now()
        sf.save()
        return
